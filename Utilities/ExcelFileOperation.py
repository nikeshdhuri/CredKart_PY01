import openpyxl


def row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def read_data(file, sheetname, rowno, colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowno, column=colno).value


def write_data(file, sheetname, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = file[sheetname]
    sheet.cell(row=rowno, column=colno).value = data
    workbook.save(file)
